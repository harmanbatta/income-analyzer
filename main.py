import os
import json
import re
import uuid
import asyncio
import tempfile
import base64
from collections import Counter
from datetime import datetime
from typing import Optional

import anthropic
import fitz  # PyMuPDF
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
if not ANTHROPIC_API_KEY:
    raise RuntimeError("ANTHROPIC_API_KEY environment variable is required")

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
sessions: dict[str, dict] = {}
jobs: dict[str, dict] = {}
_running_tasks: set = set()


class UpdateTransactionRequest(BaseModel):
    include: bool
    reason: Optional[str] = None


class BulkUpdateRequest(BaseModel):
    description: str          # substring to match (case-insensitive)
    include: bool             # value to set on all matching transactions


# ─────────────────────────────────────────────────────────────────────────────
# JSON extraction utilities (robust parsing of Claude output)
# ─────────────────────────────────────────────────────────────────────────────

def extract_balanced(text: str, open_ch: str, close_ch: str, start: int) -> tuple[int, int]:
    depth = 0
    in_string = False
    escape_next = False
    for i, ch in enumerate(text[start:], start):
        if escape_next:
            escape_next = False
            continue
        if ch == "\\" and in_string:
            escape_next = True
            continue
        if ch == '"' and not escape_next:
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch == open_ch:
            depth += 1
        elif ch == close_ch:
            depth -= 1
            if depth == 0:
                return start, i
    return start, -1


def extract_json_transactions(text: str) -> list:
    text = re.sub(r"^```[a-z]*\n?", "", text.strip(), flags=re.MULTILINE)
    text = re.sub(r"\n?```$", "", text.strip(), flags=re.MULTILINE)
    text = text.strip()
    try:
        result = json.loads(text)
        if isinstance(result, list):
            return result
        if isinstance(result, dict) and "transactions" in result:
            return result["transactions"]
    except json.JSONDecodeError:
        pass
    obj_start = text.find("{")
    if obj_start != -1:
        _, obj_end = extract_balanced(text, "{", "}", obj_start)
        if obj_end != -1:
            try:
                obj = json.loads(text[obj_start: obj_end + 1])
                if isinstance(obj, dict) and "transactions" in obj:
                    return obj["transactions"]
            except json.JSONDecodeError:
                pass
    arr_start = text.find("[")
    if arr_start == -1:
        raise ValueError("No JSON array or object found in response")
    _, arr_end = extract_balanced(text, "[", "]", arr_start)
    if arr_end != -1:
        try:
            return json.loads(text[arr_start: arr_end + 1])
        except json.JSONDecodeError:
            pass
    partial = text[arr_start:]
    last_complete = max(partial.rfind("},"), partial.rfind("}\n"))
    if last_complete != -1:
        partial = partial[: last_complete + 1] + "]"
        try:
            return json.loads(partial)
        except json.JSONDecodeError:
            pass
    raise ValueError(f"Could not parse JSON from response. Length={len(text)}")


def deduplicate_categories(transactions: list) -> list:
    """Post-processing: normalize inconsistent category/include for same description+type."""
    key_cats: dict = {}
    key_incs: dict = {}
    for tx in transactions:
        key = (tx.get("description", "").strip().lower(), tx.get("type", ""))
        if key not in key_cats:
            key_cats[key] = Counter()
            key_incs[key] = Counter()
        key_cats[key][tx.get("category", "")] += 1
        key_incs[key][tx.get("include", True)] += 1
    canonical_cat = {k: v.most_common(1)[0][0] for k, v in key_cats.items()}
    canonical_inc = {k: v.most_common(1)[0][0] for k, v in key_incs.items()}
    for tx in transactions:
        key = (tx.get("description", "").strip().lower(), tx.get("type", ""))
        tx["category"] = canonical_cat[key]
        tx["include"] = canonical_inc[key]
    return transactions


# ─────────────────────────────────────────────────────────────────────────────
# Stage 1 — Python PDF parser (no AI, deterministic)
# Works for BMO and most Canadian bank statements with month-day date format.
# ─────────────────────────────────────────────────────────────────────────────

# Date patterns — handles multiple Canadian bank formats:
#   BMO / Scotia:  "Apr 01"      (month abbr + day, no year)
#   Affinity:      "3 Mar 2025"  (day + month abbr + 4-digit year)
DATE_RE = re.compile(
    r'^(?:'
    r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2}'           # BMO/Scotia
    r'|\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{4}'  # Affinity
    r')$', re.I
)
AMOUNT_RE = re.compile(r'^\$?-?[\d,]+\.\d{2}$')   # optional $ prefix (Affinity)
PERIOD_RE = re.compile(r'for the period ending\s+\S+\s+\d{1,2},?\s+(\d{4})', re.I)
YEAR_RE   = re.compile(r'\b(20\d{2})\b')

MONTH_TO_NUM = {
    'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
    'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
    'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12',
}
NUM_TO_MONTH = {v: k.capitalize() for k, v in MONTH_TO_NUM.items()}

# When any of these lines appear, stop collecting transactions (non-transaction sections)
STOP_SECTIONS = {
    'my cheques', 'my loans', '~ end of statement ~',
    'end of statement', 'loan summary:', 'my cheques & savings account',
    'total',   # summary row — amounts that follow are section totals, not transactions
}

# Lines to always discard (matched against lowercased content)
DISCARD_EXACT = {
    # BMO
    'transaction details', 'transaction details (continued)',
    'amounts debited', 'amounts credited',
    'from your account ($)', 'to your account ($)', 'balance ($)',
    '(continued)', 'continued',
    'business banking statement', 'business banking',
    'summary of account', 'number of items processed',
    'your branch address:', 'your branch', 'your plan',
    'direct banking', 'www.bmo.com',
    'essential plan $0 monthly fee',
    'for questions about your', 'statement call',
    'debited ($)', 'credited ($)', 'balance ($) on',
    # Scotia
    'your preferred package account summary',
    'minus total withdrawals', 'plus total deposits',
    "here's what happened in your account this statement period",
    'amounts', 'transactions', 'withdrawn ($)', 'deposited ($)',
    'call 1 800 4-scotia', 'for online account access:',
    'your account number:', 'questions?',
    # Affinity
    'my chequing & savings account', 'statement of accounts',
    'statement reconciliation', 'credit union deposit insurance',
    'bill payments', 'continued...', 'my account summary',
    'deposits \u2013 cdn', 'deposits \u2013 usd', 'loans',
    'withdrawals', 'deposits', 'balance',
    'business select chequing sub number 001',
    'business savings sub number 001',
    'advances', 'payments', 'principal', 'interest',
    'current interest rate', 'interest paid current year',
    # General
    'date', 'description', 'opening', 'closing', 'account',
    '-', '+', '=', '|',
}

DISCARD_PATTERNS = [
    re.compile(r'^page \d+ of \d+$', re.I),           # "Page 1 of 17"
    re.compile(r'^\.*$'),                               # blank or all-dots
    re.compile(r'^\.{2,}'),                             # lines starting with dots
    re.compile(r'^-{3,}$'),                             # separator lines
    re.compile(r'^\d+$'),                               # bare integers (e.g. item counts 76, 516)
    re.compile(r'^business account\b', re.I),           # "Business Account # ..."
    re.compile(r'^business name:', re.I),
    re.compile(r'^#\s*\d'),                             # "# 3050 1981-377"
    re.compile(r'^transit number:', re.I),
    re.compile(r'^\(\d{3}\)\s*\d{3}'),                 # phone numbers
    re.compile(r'^1-\d{3}-\d{3}-\d{4}$'),              # 1-800 numbers
    re.compile(r'^www\.', re.I),                        # URLs
    re.compile(r'^[a-z]\d[a-z]\s*\d[a-z]\d$', re.I),  # Canadian postal codes like L6R3P4
    re.compile(r'^\d{5}\s+[A-Z]{2}'),                  # US-style zip + state
    re.compile(r'^for the period ending', re.I),        # period header line
]

# Date-prefixed lines that are NOT real transactions
NON_TX_DESCRIPTIONS = {
    'opening balance', 'closing totals', 'closing balance', 'balance forward',
}


def _should_discard(line: str) -> bool:
    low = line.lower()
    if low in DISCARD_EXACT:
        return True
    return any(p.match(low) for p in DISCARD_PATTERNS)


def _parse_amount(s: str) -> float:
    return float(s.replace(',', '').replace('$', ''))


def _extract_account_holder(pdf_bytes: bytes) -> str:
    """
    Pull the account holder / business name from the first page of any PDF.
    Used to detect internal same-name e-transfers (savings ↔ chequing).
    Returns empty string if not found.
    """
    try:
        doc  = fitz.open(stream=pdf_bytes, filetype='pdf')
        text = doc[0].get_text() if len(doc) > 0 else ""
        doc.close()
    except Exception:
        return ""

    # BMO: "Business name:\n<NAME>"
    m = re.search(r'Business name:\s*\n?\s*([A-Z][A-Z0-9 &.,\'-]{3,})', text, re.I)
    if m:
        return m.group(1).strip()

    # Affinity / generic: "Account Number XXXX - <NAME>"
    m = re.search(r'Account Number\s+\S+\s*[-–]\s*(.+?)(?:\n|$)', text, re.I)
    if m:
        return m.group(1).strip()

    # Scotia / RBC style: bold name block near top (all-caps 2+ words)
    m = re.search(r'^([A-Z]{2,}(?:\s+[A-Z]{2,}){1,5})\s*$', text, re.M)
    if m:
        candidate = m.group(1).strip()
        # Skip obvious non-names
        skip = {'ACCOUNT STATEMENT', 'TRANSACTION DETAILS', 'ROYAL BANK', 'TD CANADA TRUST',
                'CIBC', 'SCOTIABANK', 'BMO', 'AFFINITY', 'WEALTHSIMPLE', 'BALANCE FORWARD'}
        if candidate not in skip and len(candidate) > 5:
            return candidate

    return ""


def extract_transactions_from_pdf(pdf_bytes: bytes, filename: str) -> list[dict]:
    """
    Pure Python extraction from a bank statement PDF.

    PyMuPDF linearises the table layout so each field (date, description,
    continuation lines, amount, balance) appears on its own text line.

    Algorithm:
      1. Collect all non-garbage lines from every page.
      2. Group into blocks, each starting with a date line (Mon DD).
      3. Within each block, the last two numbers are amount + balance.
         Everything in between is the transaction description.
      4. Determine credit/debit from balance delta — no guessing from text.
    """
    all_lines = []
    raw_pages = []

    try:
        doc = fitz.open(stream=pdf_bytes, filetype='pdf')
        for page in doc:
            page_text = page.get_text()
            raw_pages.append(page_text)
            for line in page_text.split('\n'):
                line = line.strip()
                if line and not _should_discard(line):
                    all_lines.append(line)
        doc.close()
    except Exception as e:
        print(f"[WARN] fitz error for {filename}: {e}")
        return []

    full_text = '\n'.join(raw_pages)

    # Detect statement year from "For the period ending … YYYY"
    year = datetime.now().year
    m = PERIOD_RE.search(full_text)
    if m:
        year = int(m.group(1))
    else:
        found = YEAR_RE.findall(full_text[:800])
        if found:
            year = int(found[0])

    print(f"[INFO] {filename}: year={year}, {len(all_lines)} lines after filtering")

    # Group lines into blocks — each block starts with a date line.
    # Stop collecting when a section-terminator line is encountered
    # (e.g. "my cheques", "my loans") to avoid double-counting.
    blocks = []
    current: list[str] = []
    stopped = False
    for line in all_lines:
        if line.lower() in STOP_SECTIONS:
            stopped = True
            if current:
                blocks.append(current)
                current = []
            break
        if DATE_RE.match(line):
            if current:
                blocks.append(current)
            current = [line]
        elif current:
            current.append(line)
    if current and not stopped:
        blocks.append(current)

    print(f"[INFO] {filename}: {len(blocks)} date-prefixed blocks found"
          + (" (stopped at section terminator)" if stopped else ""))

    transactions = []
    prev_balance: float | None = None

    for block in blocks:
        date_line = block[0]
        rest      = block[1:]

        # Pull trailing numbers off the end: second-to-last = amount, last = balance
        trailing: list[str] = []
        desc_parts: list[str] = []
        for item in reversed(rest):
            if AMOUNT_RE.match(item) and len(trailing) < 2:
                trailing.insert(0, item)
            else:
                desc_parts.insert(0, item)

        description = ' '.join(desc_parts).strip()

        # Only 1 trailing number → opening balance row (no transaction amount)
        if len(trailing) < 2:
            if trailing:
                try:
                    prev_balance = _parse_amount(trailing[0])
                except ValueError:
                    pass
            continue

        # Skip known non-transaction date rows
        if description.lower() in NON_TX_DESCRIPTIONS:
            continue

        amount_str, balance_str = trailing[0], trailing[1]
        try:
            amount  = _parse_amount(amount_str)
            balance = _parse_amount(balance_str)
        except ValueError:
            continue

        # Determine direction from balance delta — 100% accurate, no description guessing
        direction = 'debit'  # safe default
        if prev_balance is not None:
            delta = round(balance - prev_balance, 2)
            if abs(delta - amount) < 0.02:      # balance went up → credit
                direction = 'credit'
            elif abs(delta + amount) < 0.02:    # balance went down → debit
                direction = 'debit'
            else:
                # Delta mismatch — fall back to description keywords
                dl = description.lower()
                if any(k in dl for k in ['received', 'direct deposit', 'deposit', 'refund']):
                    direction = 'credit'
                elif any(k in dl for k in ['sent', 'purchase', 'payment', 'withdrawal',
                                            'fee', 'charge', 'draft', 'transfer out']):
                    direction = 'debit'
                else:
                    direction = 'debit'

        # Build date and month strings — handle two formats:
        #   BMO/Scotia: "Apr 01"      → month=parts[0], day=parts[1], year from statement
        #   Affinity:   "3 Mar 2025"  → day=parts[0], month=parts[1], year=parts[2]
        parts = date_line.split()
        if len(parts) == 3 and parts[2].isdigit() and len(parts[2]) == 4:
            # Affinity format — year is embedded in the date line itself
            m_abbr  = parts[1].capitalize()
            day     = parts[0].zfill(2)
            tx_year = int(parts[2])          # always correct, even across year boundaries
        else:
            # BMO / Scotia format
            m_abbr  = parts[0].capitalize()
            day     = parts[1].zfill(2)
            tx_year = year

        mon_num     = MONTH_TO_NUM.get(m_abbr.lower(), '01')
        full_date   = f"{tx_year}-{mon_num}-{day}"
        month_field = f"{m_abbr}-{str(tx_year)[2:]}"   # e.g. "Mar-25"

        transactions.append({
            'id':          str(uuid.uuid4()),
            'date':        full_date,
            'description': description,
            'amount':      amount,
            'direction':   direction,
            'month':       month_field,
        })
        prev_balance = balance

    credits = sum(1 for t in transactions if t['direction'] == 'credit')
    debits  = sum(1 for t in transactions if t['direction'] == 'debit')
    print(f"[INFO] {filename}: extracted {len(transactions)} transactions "
          f"({credits} credits / {debits} debits)")
    return transactions


# ─────────────────────────────────────────────────────────────────────────────
# Stage 2 — Claude categorisation prompt (categorisation only, no extraction)
# ─────────────────────────────────────────────────────────────────────────────

CAT_PROMPT = """MORTGAGE UNDERWRITING — TRANSACTION CATEGORISATION

You are a mortgage underwriting analyst. Bank transactions have been pre-extracted by a parser. Your only job is to assign category, suggested_include (Y or N), and reason to each transaction.

CRITICAL: The "direction" field is pre-determined from the bank statement and is always correct.
  direction "credit" = money coming IN  (income transaction)
  direction "debit"  = money going OUT  (expense transaction)
Never override the direction. Never reclassify a debit as income or a credit as expense.

Return ONLY a valid JSON array — no prose, no markdown fences.
Each element: {"id": "...", "category": "...", "suggested_include": "Y", "reason": "..."}
Every input transaction must appear in the output with its exact id.

── CATEGORIES ──────────────────────────────────────────────────────────────
Use the most specific label from the payee name or transfer type.
Income: INTERAC e-Transfer In, ATM Deposit, Mobile Deposit, Cheque Deposit, Direct Deposit, Wire Transfer, Cash Deposit, Internal Transfer In, NSF Reversal, Government Rebate, Merchant Services Deposit, or exact sender name.
Expense: exact payee name (e.g. "Enbridge Gas", "CHIT CHATS BC") or transfer type such as INTERAC e-Transfer Out, Cash Withdrawal, Bank Charges, Cheque, CC Transfer, Transfer Out, Canadian Draft, Merchant Services Fee.
Special: MSP fees / merchant services fees / POS fees / terminal fees → always "Merchant Services Fee", always Y.
Use "Other Income" or "Other Expense" only if truly unclear; explain in reason.

── MERCHANT SERVICES DEPOSITS (critical rule) ──────────────────────────────
ROYAL BANK CENTRAL CARD CENTRE, MONERIS, SQUARE, STRIPE, ELAVON, CHASE PAYMENTECH,
GLOBAL PAYMENTS, TD MERCHANT SERVICES, and any similar payment processor name:
  • direction = credit (money IN): This is a merchant services deposit — daily credit card
    terminal settlement deposited into the business account. Category: "Merchant Services Deposit",
    suggested_include: Y. NEVER classify these as credit card payments — they are INCOME.
  • direction = debit (money OUT): This is a chargeback, fee, or reversal from the processor.
    Category: "Merchant Services Fee", suggested_include: Y — these are legitimate business expenses.
This rule is absolute — a deposit from ROYAL BANK CENTRAL CARD CENTRE is always income.

── CONSISTENCY RULE (absolute) ─────────────────────────────────────────────
Identical or near-identical descriptions must always get identical category and identical suggested_include.
Before finalising, scan all transactions and fix any conflicts.

── INCOME RULES (direction = credit) ───────────────────────────────────────

Auto-set N:
- Internal own-account transfers: keywords TFR-FR, transfer from own, same account holder → reason: Internal transfer — not external income.
- SAME-NAME TRANSFERS (critical): If an e-transfer is received FROM a name that matches or closely resembles the account holder's own business or personal name, it is an internal transfer between the account holder's own accounts (e.g. chequing to savings). Set N, category "Internal Transfer", reason "Internal transfer between own accounts — excluded." Apply this even if the account holder name is not explicitly provided — look for the name that appears on the statement header.
- NSF reversals and re-credits → reason: NSF reversal — not new income.
- Government rebates: HST rebate, GST credit, Ontario Trillium Benefit → reason: Government rebate — excluded.
- Wire transfers (any amount) → reason: Review: wire transfer — verify source and whether qualifying income.
- Loan or LOC deposits: keywords loan, LOC, credit line, financing, advance, lendcare, or any lender name → reason: Review: possible loan deposit — verify if qualifying income.
- Single deposit significantly larger than average deposits in this batch → reason: Review: unusually large deposit — verify source.
- Returned outgoing e-transfers (sent out then returned) → reason: Returned outgoing e-transfer — not new income.

Auto-set Y:
- INTERAC e-transfers from clearly external senders.
- ATM / mobile / cash / cheque deposits from third parties.
- Regular recurring deposits that appear employment or business-related.
- Direct deposits from known employers or processors (Square, Stripe, PayPal, etc.).
- When uncertain, default Y and note in reason.

── EXPENSE RULES (direction = debit) ───────────────────────────────────────

Auto-set N — never override these rules:
- ALL bank fees without exception: monthly plan fees, NSF fees, overdraft charges, overdraft per item charges, e-transfer fees, wire fees, service charges, transaction fees, draft fees, excess item fees → category: Bank Charges, reason: Bank fee — excluded.
- Credit card payments: keywords MC, VISA, AMEX, MASTERCARD, CAN TIRE MC, TD VISA, CIBC VISA, RBC VISA, SCOTIA VISA, BMO MC, or any description combining a card brand with alphanumeric characters → reason: Credit card payment — excluded.
- Transfers to own accounts at same or other institutions → reason: Internal transfer — excluded.
- SAME-NAME TRANSFERS (critical): If an e-transfer is sent TO a name that matches or closely resembles the account holder's own business or personal name, it is an internal transfer to the account holder's own savings or other account. Set N, category "Internal Transfer", reason "Internal transfer between own accounts — excluded."
- CRA, Receiver General, Canada Revenue Agency, or any government tax remittance: keywords CRA, CANADA REVENUE, RECEIVER GENERAL, HST REMIT, GST REMIT, TAX REMIT → category: CRA / Tax Remittance, reason: CRA / government tax remittance — excluded from qualifying expenses. This rule is absolute — never set Y for CRA or Receiver General payments.

Auto-set Y:
- Insurance payments.
- Utilities: gas, hydro, internet, cable, phone, electricity.
- Loan and mortgage payments to external lenders.
- INTERAC e-transfers out to clearly external payees.
- Cheques to third-party individuals or businesses.
- Rent payments.
- Debit card purchases at identifiable merchants.
- Any regular, identifiable recurring expense.
- When uncertain, default Y.

── FLAGGING ────────────────────────────────────────────────────────────────
Unusual, one-time, very large, or unclear transactions → start reason with "Review:" followed by concern.
"""


# ─────────────────────────────────────────────────────────────────────────────
# Stage 1b — Claude vision extraction (image-based PDFs)
# Used when PyMuPDF returns no text (scanned PDFs, image-rendered tables, etc.)
# Combines extraction + categorisation in one call per page batch.
# ─────────────────────────────────────────────────────────────────────────────

IMAGE_EXTRACT_PROMPT = """You are a mortgage underwriting analyst reading bank statement images.

Extract EVERY transaction visible in these images and return a JSON array.
Skip only: opening balance rows, closing balance rows, total/subtotal rows, balance forward rows.
IMPORTANT: If you see pages that show physical cheque images (handwritten or printed cheques),
skip those pages entirely — do not extract amounts from cheque images.
Only extract from pages showing a transaction table with Date | Description | Withdrawals | Deposits | Balance columns.
Include every actual debit and credit — do not skip any from the transaction table.

Each element of the array must have ALL these fields:
{
  "date": "YYYY-MM-DD",
  "description": "exact description text from statement",
  "amount": 123.45,
  "direction": "credit" or "debit",
  "month": "Mon-YY",
  "category": "...",
  "suggested_include": "Y" or "N",
  "reason": "brief reason, max 12 words"
}

DIRECTION: Look at which column the amount appears in.
  Deposits / Credits column → "credit"  (money coming in)
  Withdrawals / Debits column → "debit"  (money going out)
amount is always a positive number regardless of direction.
month format example: Apr-24, Jan-25, Mar-26.

CATEGORIES: Use specific labels. Income: INTERAC e-Transfer In, Direct Deposit, Payroll Deposit,
ATM Deposit, Cheque Deposit, Merchant Services Deposit, Wire Transfer In, or exact sender name.
Expense: exact payee name or type such as INTERAC e-Transfer Out, Cheque, Mortgage Payment,
Loan Payment, CC Transfer, Cash Withdrawal, Bank Charges, Merchant Services Fee.

INCOME — Auto N: own-account internal transfers, wire transfers (add "Review:" prefix),
government rebates (HST/GST credit, OTB), loan/LOC deposits, returned e-transfers, unusually large deposits.
INCOME — Auto Y: external e-transfers, payroll deposits, direct deposits, merchant card settlements,
ATM/cash/cheque deposits from third parties.

EXPENSE — Auto N (never override): ALL bank fees (monthly fees, NSF, overdraft, e-transfer fees,
service charges, cheque image fees), credit card payments (VISA/MC/AMEX/Mastercard bill payments),
own-account transfers, CRA/Receiver General/government tax payments (PAD CCRA, CRA Source Deduct, etc.).
EXPENSE — Auto Y: insurance, utilities, loan payments, rent, identifiable merchant purchases,
e-transfers out to external payees, cheques to third parties.

Return ONLY valid JSON array. No prose, no markdown.
"""


async def _extract_via_vision(pdf_bytes: bytes, filename: str,
                              extra_context: str = "") -> list[dict]:
    """
    Fallback for image-based PDFs: render each page as PNG and send to Claude
    vision for combined extraction + categorisation.
    Returns fully-formed transaction dicts (already have category/include/reason).
    """
    PAGES_PER_BATCH = 4
    RENDER_SCALE    = 1.8   # zoom for readable resolution without huge file size

    try:
        doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    except Exception as e:
        print(f"[WARN] Cannot open {filename} for vision: {e}")
        return []

    # Detect year from first page text (may still have text in header)
    year = datetime.now().year
    first_text = doc[0].get_text() if len(doc) > 0 else ""
    m = YEAR_RE.search(first_text)
    if m:
        year = int(m.group(1))

    # Render all pages
    page_images: list[tuple[int, bytes]] = []
    matrix = fitz.Matrix(RENDER_SCALE, RENDER_SCALE)
    for page_num, page in enumerate(doc, 1):
        try:
            pix      = page.get_pixmap(matrix=matrix)
            img_data = pix.tobytes("png")
            page_images.append((page_num, img_data))
        except Exception as e:
            print(f"[WARN] Could not render page {page_num} of {filename}: {e}")
    doc.close()

    if not page_images:
        print(f"[WARN] {filename}: no pages could be rendered")
        return []

    print(f"[INFO] {filename}: image-based PDF, {len(page_images)} pages → vision extraction")

    all_transactions: list[dict] = []

    for batch_start in range(0, len(page_images), PAGES_PER_BATCH):
        batch = page_images[batch_start: batch_start + PAGES_PER_BATCH]

        # Build multimodal message content
        content: list[dict] = []
        for page_num, img_bytes in batch:
            content.append({"type": "text", "text": f"Page {page_num}:"})
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/png",
                    "data": base64.b64encode(img_bytes).decode(),
                },
            })
        content.append({"type": "text", "text": IMAGE_EXTRACT_PROMPT + extra_context})

        def _vision_call(c=content):
            return client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=16000,
                messages=[{"role": "user", "content": c}],
            ).content[0].text.strip()

        try:
            response_text = await asyncio.to_thread(_vision_call)
            batch_txs     = extract_json_transactions(response_text)
            print(f"[INFO] {filename} vision batch pages "
                  f"{batch[0][0]}-{batch[-1][0]}: {len(batch_txs)} transactions")

            for tx in batch_txs:
                # Normalise and fill required fields
                direction = tx.get("direction", "debit").lower()
                tx_type   = "income" if direction == "credit" else "expense"
                amount    = abs(float(tx.get("amount", 0)))
                signed    = amount if tx_type == "income" else -amount
                inc_flag  = str(tx.get("suggested_include", "Y")).strip().upper() == "Y"

                # Fix date year if Claude guessed wrong
                date_str = tx.get("date", f"{year}-01-01")
                if len(date_str) >= 4 and date_str[:4].isdigit():
                    pass  # year already present
                else:
                    date_str = f"{year}-{date_str}"

                # Fix month field
                month_str = tx.get("month", "")
                if not month_str:
                    try:
                        dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
                        month_str = dt.strftime("%b-%y")
                    except Exception:
                        month_str = f"Unk-{str(year)[2:]}"

                all_transactions.append({
                    "id":               str(uuid.uuid4()),
                    "date":             date_str[:10],
                    "description":      tx.get("description", ""),
                    "amount":           signed,
                    "type":             tx_type,
                    "month":            month_str,
                    "category":         tx.get("category", "Other"),
                    "suggested_include": tx.get("suggested_include", "Y"),
                    "reason":           tx.get("reason", ""),
                    "include":          inc_flag,
                    "_vision":          True,   # flag: already categorised
                })

        except Exception as e:
            print(f"[WARN] Vision extraction failed for {filename} "
                  f"pages {batch[0][0]}-{batch[-1][0]}: {e}")

    print(f"[INFO] {filename}: vision complete — {len(all_transactions)} transactions")
    return all_transactions


# ─────────────────────────────────────────────────────────────────────────────
# Frontend HTML + JavaScript
# ─────────────────────────────────────────────────────────────────────────────

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Sahara Capital &mdash; Income Analyzer</title>
  <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,700;1,700&family=DM+Sans:wght@400;500;600&display=swap" rel="stylesheet">
  <script src="https://cdn.tailwindcss.com"></script>
  <style>
    *  { box-sizing: border-box; }
    @keyframes spin { to { transform: rotate(360deg); } }
    .spin { animation: spin 1s linear infinite; }
    .clamp2 { overflow:hidden; display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; }
    body { font-family:'DM Sans',system-ui,sans-serif; background:#e8e4dc; color:#1a1a1a; -webkit-font-smoothing:antialiased; }
    .sc-serif   { font-family:'Playfair Display',Georgia,serif; font-weight:700; }
    .sc-badge   { display:inline-block; border:1px solid #c4a050; color:#c4a050; font-size:10px; font-weight:600; letter-spacing:2.5px; text-transform:uppercase; padding:6px 14px; }
    .sc-hero    { background:#1B3D2C; padding:40px 48px 44px; }
    .sc-hero-sm { background:#1B3D2C; padding:16px 28px; }
    .sc-strip   { background:#152E21; display:flex; width:100%; }
    .sc-strip-cell { flex:1; text-align:center; padding:20px 8px; border-right:1px solid rgba(255,255,255,0.1); }
    .sc-strip-cell:last-child { border-right:none; }
    .sc-strip-num  { font-family:'Playfair Display',Georgia,serif; font-size:26px; font-weight:700; color:#c4a050; line-height:1; }
    .sc-strip-lbl  { font-size:9px; font-weight:600; letter-spacing:1.5px; text-transform:uppercase; color:rgba(255,255,255,0.55); margin-top:5px; }
    .sc-divider { width:44px; height:3px; background:#c4a050; }
    .sc-card    { background:#fff; border-left:4px solid #1B3D2C; }
    .sc-btn     { background:#1B3D2C; color:#fff; font-size:12px; font-weight:700; letter-spacing:1.5px; text-transform:uppercase; padding:14px 28px; display:inline-flex; align-items:center; gap:8px; cursor:pointer; border:none; transition:opacity .15s; }
    .sc-btn:hover { opacity:.88; }
    .sc-btn:disabled { opacity:.45; cursor:not-allowed; }
    .sc-btn-gold { background:#c4a050; }
    .sc-btn-full { width:100%; justify-content:center; }
    .sc-contact { background:#1B3D2C; display:flex; width:100%; }
    .sc-contact-info { flex:1; padding:32px 36px; }
    .sc-contact-cta  { width:38%; background:#c4a050; padding:32px 24px; text-align:center; }
    @media(max-width:640px) {
      .sc-hero { padding:28px 22px 32px; }
      .sc-strip { flex-wrap:wrap; }
      .sc-strip-cell { width:50%; border-bottom:1px solid rgba(255,255,255,0.1); }
      .sc-strip-cell:nth-child(1),.sc-strip-cell:nth-child(3) { border-right:1px solid rgba(255,255,255,0.1); }
      .sc-contact { flex-direction:column; }
      .sc-contact-cta { width:100%; }
    }
  </style>
</head>
<body class="m-0 p-0">
<div id="app"></div>
<script>
(function() {

  var state = {
    page: 'upload',
    files: [],
    isDragging: false,
    isAnalyzing: false,
    analyzeError: null,
    jobId: null,
    progressInfo: {},
    sessionId: null,
    transactions: [],
    filter: 'all',
    search: '',
    updating: {},
    downloading: false,
    downloaded: false
  };

  var pollTimer = null;
  var dragCounter = 0;   // fix: counter prevents jitter from child-element drag events

  function h(str) {
    return String(str)
      .replace(/&/g,'&amp;').replace(/</g,'&lt;')
      .replace(/>/g,'&gt;').replace(/"/g,'&quot;');
  }

  function fmt(n) {
    return new Intl.NumberFormat('en-US',{style:'currency',currency:'USD'}).format(n);
  }

  function getFiltered() {
    var r = state.transactions.slice();
    if (state.filter === 'income')    r = r.filter(function(t){return t.type==='income';});
    else if (state.filter === 'expense')  r = r.filter(function(t){return t.type==='expense';});
    else if (state.filter === 'included') r = r.filter(function(t){return t.include;});
    else if (state.filter === 'excluded') r = r.filter(function(t){return !t.include;});
    if (state.search.trim()) {
      var q = state.search.toLowerCase();
      r = r.filter(function(t){
        return t.description.toLowerCase().indexOf(q)>=0 ||
               t.category.toLowerCase().indexOf(q)>=0 ||
               t.date.indexOf(q)>=0 ||
               t.reason.toLowerCase().indexOf(q)>=0;
      });
    }
    return r;
  }

  function getTotals() {
    var inc = state.transactions.filter(function(t){return t.include;});
    var income   = inc.filter(function(t){return t.type==='income';}).reduce(function(s,t){return s+t.amount;},0);
    var expenses = inc.filter(function(t){return t.type==='expense';}).reduce(function(s,t){return s+Math.abs(t.amount);},0);
    return {income:income, expenses:expenses, net:income-expenses,
            count:state.transactions.length, includedCount:inc.length};
  }

  function setState(updates) {
    Object.assign(state, updates);
    render();
  }

  function render() {
    var activeId = document.activeElement ? document.activeElement.id : null;
    var scrollY  = window.scrollY;
    var app = document.getElementById('app');
    if      (state.page === 'upload')   app.innerHTML = renderUpload();
    else if (state.page === 'progress') app.innerHTML = renderProgress();
    else if (state.page === 'review')   app.innerHTML = renderReview();
    else                                app.innerHTML = renderDownload();
    attachListeners();
    if (activeId) {
      var el = document.getElementById(activeId);
      if (el) {
        el.focus();
        if ((el.tagName==='INPUT'||el.tagName==='TEXTAREA') && el.type!=='file') {
          try { el.setSelectionRange(el.value.length, el.value.length); } catch(e){}
        }
      }
    }
    window.scrollTo(0, scrollY);
  }

  /* ============================================================
     UPLOAD PAGE
  ============================================================ */
  function renderUpload() {
    var fileRows = state.files.map(function(f,i){
      return '<div class="flex items-center justify-between bg-white border border-gray-200 rounded-xl px-4 py-3 shadow-sm">' +
        '<div class="flex items-center gap-3 min-w-0">' +
          '<div class="w-8 h-8 rounded-lg bg-red-50 flex items-center justify-center flex-shrink-0 text-red-500 text-xs font-bold">PDF</div>' +
          '<div class="min-w-0">' +
            '<p class="text-sm font-medium text-gray-800 truncate">' + h(f.name) + '</p>' +
            '<p class="text-xs text-gray-400">' + (f.size/1024).toFixed(1) + ' KB</p>' +
          '</div>' +
        '</div>' +
        '<button data-remove="' + i + '" class="ml-3 p-1.5 rounded-lg hover:bg-red-50 text-gray-300 hover:text-red-500 transition-colors text-lg leading-none">&times;</button>' +
      '</div>';
    }).join('');

    var dz = state.isDragging
      ? 'border-[#1B3D2C] bg-[#f5f2ec]'
      : 'border-gray-300 bg-white hover:border-[#1B3D2C]';
    var btnDis = state.isAnalyzing || state.files.length === 0;
    var btnCls = btnDis
      ? 'bg-gray-200 text-gray-400 cursor-not-allowed'
      : 'sc-bg text-white hover:opacity-90 shadow-md hover:shadow-lg active:scale-95';

    return (
      // ── Hero header ────────────────────────────────────────────────
      '<div class="sc-hero">' +
        '<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:30px">' +
          '<div class="sc-badge">Income Calculation Tool</div>' +
          '<img src="https://saharacapital.ca/wp-content/uploads/2026/01/Logo-02.png" alt="Sahara Capital" style="height:52px;width:auto">' +
        '</div>' +
        '<div style="font-size:10px;font-weight:600;letter-spacing:2px;color:rgba(255,255,255,0.5);text-transform:uppercase;margin-bottom:10px">Team Harman Batta</div>' +
        '<div class="sc-serif" style="font-size:42px;color:#ffffff;line-height:1.15">Sahara Capital Group</div>' +
        '<div class="sc-serif" style="font-size:26px;color:#c4a050;font-style:italic;margin-top:6px">by your side, always</div>' +
        '<div class="sc-divider" style="margin:20px 0"></div>' +
        '<div style="font-size:19px;color:rgba(255,255,255,0.9);font-weight:600">Income Analyzer</div>' +
        '<div style="font-size:14px;color:rgba(255,255,255,0.5);margin-top:6px">Upload PDF bank statements for mortgage underwriting analysis</div>' +
      '</div>' +

      // ── Body ───────────────────────────────────────────────────────
      '<div style="background:#e8e4dc;padding:32px;min-height:60vh">' +
        '<div style="max-width:580px;margin:0 auto">' +

          // File rows
          (state.files.length > 0
            ? '<div style="margin-bottom:16px;display:flex;flex-direction:column;gap:8px">' +
                state.files.map(function(f,i){
                  return '<div style="display:flex;align-items:center;justify-content:space-between;background:#fff;border-left:4px solid #1B3D2C;padding:12px 16px">' +
                    '<div style="display:flex;align-items:center;gap:12px;min-width:0">' +
                      '<div style="background:#1B3D2C;color:#fff;font-size:9px;font-weight:700;padding:4px 7px;letter-spacing:1px">PDF</div>' +
                      '<div style="min-width:0">' +
                        '<p style="font-size:13px;font-weight:600;color:#1a1a1a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">' + h(f.name) + '</p>' +
                        '<p style="font-size:11px;color:#999">' + (f.size/1024).toFixed(1) + ' KB</p>' +
                      '</div>' +
                    '</div>' +
                    '<button data-remove="' + i + '" style="background:none;border:none;color:#999;font-size:20px;cursor:pointer;padding:4px 8px;line-height:1">&times;</button>' +
                  '</div>';
                }).join('') +
              '</div>'
            : '') +

          // Error
          (state.analyzeError
            ? '<div style="background:#fff0f0;border-left:4px solid #c0392b;padding:12px 16px;margin-bottom:16px;font-size:13px;color:#c0392b">' + h(state.analyzeError) + '</div>'
            : '') +

          // Drop zone
          '<input id="file-input" type="file" accept=".pdf,application/pdf" multiple style="display:none">' +
          '<div id="drop-zone" style="border:2px dashed ' + (state.isDragging ? '#c4a050' : '#c5bfb5') + ';background:' + (state.isDragging ? '#f5f2ec' : '#fff') + ';padding:48px 24px;text-align:center;transition:all .15s;cursor:pointer;margin-bottom:20px">' +
            '<div style="margin-bottom:16px">' +
              '<svg style="width:36px;height:36px;color:' + (state.isDragging ? '#1B3D2C' : '#a0998f') + ';display:inline-block" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="1.5" d="M7 16a4 4 0 01-.88-7.903A5 5 0 1115.9 6L16 6a5 5 0 011 9.9M15 13l-3-3m0 0l-3 3m3-3v12"/></svg>' +
            '</div>' +
            '<p style="font-weight:600;color:#1B3D2C;font-size:15px;margin-bottom:6px">' + (state.isDragging ? 'Drop your PDFs here' : 'Drag &amp; drop PDFs here') + '</p>' +
            '<p style="font-size:12px;color:#999;margin-bottom:20px">or use the button below to browse</p>' +
            '<label for="file-input" class="sc-btn" style="font-size:11px;padding:11px 24px">' +
              '<svg style="width:14px;height:14px" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M15.172 7l-6.586 6.586a2 2 0 102.828 2.828l6.414-6.586a4 4 0 00-5.656-5.656l-6.415 6.585a6 6 0 108.486 8.486L20.5 13"/></svg>' +
              'Choose PDF Files' +
            '</label>' +
            '<p style="font-size:11px;color:#bbb;margin-top:14px">Multiple files supported &mdash; one per month</p>' +
          '</div>' +

          // Analyze button
          '<button id="analyze-btn" ' + (state.isAnalyzing || state.files.length === 0 ? 'disabled' : '') + ' class="sc-btn sc-btn-full" style="font-size:13px;padding:16px 28px;' + (state.isAnalyzing || state.files.length === 0 ? 'opacity:.4;cursor:not-allowed' : '') + '">' +
            (state.isAnalyzing
              ? '<svg class="spin" style="width:18px;height:18px" viewBox="0 0 24 24" fill="none"><circle style="opacity:.25" cx="12" cy="12" r="10" stroke="currentColor" stroke-width="4"/><path style="opacity:.75" fill="currentColor" d="M4 12a8 8 0 018-8V0C5.373 0 0 5.373 0 12h4z"/></svg>Uploading files\u2026'
              : '<svg style="width:18px;height:18px" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12h6m-6 4h6m2 5H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>Analyze Statements') +
          '</button>' +

        '</div>' +
      '</div>' +

      // ── Legal strip ────────────────────────────────────────────────
      '<div style="background:#f5f2ec;padding:12px 32px;text-align:center">' +
        '<p style="font-size:10.5px;color:#999">Sahara Capital Group &nbsp;|&nbsp; Mortgage Alliance &nbsp;|&nbsp; FSRA Lic. #10530 &nbsp;|&nbsp; saharacapital.ca</p>' +
      '</div>'
    );
  }

  /* ============================================================
     PROGRESS PAGE
  ============================================================ */
  function renderProgress() {
    var info    = state.progressInfo || {};
    var isError = info.status === 'error';
    var pct     = (info.total_pages > 0) ? Math.round(info.pages_done / info.total_pages * 100) : 0;
    var spinnerBg = isError ? '#c0392b' : '#1B3D2C';
    var heroHtml =
      '<div class="sc-hero-sm" style="display:flex;justify-content:space-between;align-items:center">' +
        '<div>' +
          '<div style="font-size:10px;letter-spacing:2px;color:rgba(255,255,255,0.5);text-transform:uppercase">Team Harman Batta</div>' +
          '<div class="sc-serif" style="font-size:22px;color:#fff;line-height:1.2">Sahara Capital Group</div>' +
          '<div class="sc-serif" style="font-size:14px;color:#c4a050;font-style:italic">by your side, always</div>' +
        '</div>' +
        '<img src="https://saharacapital.ca/wp-content/uploads/2026/01/Logo-02.png" alt="Sahara Capital" style="height:44px;width:auto">' +
      '</div>';
    var bodyHtml =
      '<div style="background:#e8e4dc;min-height:70vh;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:40px 24px">' +
        '<div style="max-width:520px;width:100%">' +
          '<div style="text-align:center;margin-bottom:28px">' +
            '<div style="width:60px;height:60px;border-radius:50%;background:' + spinnerBg + ';display:flex;align-items:center;justify-content:center;margin:0 auto 16px">' +
              (isError
                ? '<svg style="width:28px;height:28px;color:#fff" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 9v2m0 4h.01m-6.938 4h13.856c1.54 0 2.502-1.667 1.732-3L13.732 4c-.77-1.333-2.694-1.333-3.464 0L3.34 16c-.77 1.333.192 3 1.732 3z"/></svg>'
                : '<svg class="spin" style="width:28px;height:28px;color:#fff" viewBox="0 0 24 24" fill="none"><circle style="opacity:.25" cx="12" cy="12" r="10" stroke="currentColor" stroke-width="4"/><path style="opacity:.75" fill="currentColor" d="M4 12a8 8 0 018-8V0C5.373 0 0 5.373 0 12h4z"/></svg>') +
            '</div>' +
            '<div class="sc-serif" style="font-size:28px;color:#1B3D2C">' + (isError ? 'Analysis Failed' : 'Analyzing Statements') + '</div>' +
            '<div style="font-size:13px;color:#888;margin-top:6px">' + (isError ? 'An error occurred during processing.' : 'Processing your bank statements\u2026') + '</div>' +
          '</div>' +
          '<div class="sc-card" style="background:#fff;padding:24px 28px;margin-bottom:20px">' +
            '<p style="font-size:13px;font-weight:600;color:#1B3D2C;margin-bottom:16px">' + h(info.current_file || 'Starting\u2026') + '</p>' +
            (info.total_pages > 0
              ? '<div><div style="display:flex;justify-content:space-between;font-size:11px;color:#999;margin-bottom:8px"><span>Batches categorised</span><span>' + info.pages_done + ' / ' + info.total_pages + '</span></div><div style="width:100%;background:#e8e4dc;height:6px"><div style="background:#1B3D2C;height:6px;width:' + pct + '%;transition:width .5s"></div></div><div style="font-size:11px;color:#c4a050;margin-top:8px;font-weight:600">' + pct + '% complete</div></div>'
              : '<div style="display:flex;align-items:center;gap:8px;font-size:13px;color:#999"><div style="width:8px;height:8px;background:#1B3D2C;border-radius:50%;animation:spin 1.5s linear infinite"></div><span>Extracting transactions\u2026</span></div>') +
            (isError ? '<p style="margin-top:16px;font-size:13px;color:#c0392b;background:#fff0f0;padding:12px 16px">' + h(info.error || 'Unknown error') + '</p>' : '') +
          '</div>' +
          (isError
            ? '<button id="retry-btn" class="sc-btn sc-btn-full" style="font-size:13px;padding:16px">\u2190 Back to Upload</button>'
            : '<p style="text-align:center;font-size:12px;color:#aaa">Checking progress every 5 seconds \u2014 do not close this tab.</p>') +
        '</div>' +
      '</div>';
    return heroHtml + bodyHtml;
  }

  /* ============================================================
     REVIEW PAGE
  ============================================================ */
  function renderReview() {
    var totals   = getTotals();
    var filtered = getFiltered();

    var filterBtns = ['all','income','expense','included','excluded'].map(function(v){
      var labels = {all:'All',income:'Income',expense:'Expenses',included:'Y Only',excluded:'N Only'};
      var active = state.filter === v;
      return '<button data-filter="' + v + '" style="padding:6px 14px;font-size:12px;font-weight:600;border:none;cursor:pointer;transition:all .15s;' + (active ? 'background:#1B3D2C;color:#fff' : 'background:#ede9e1;color:#666') + '">' + labels[v] + '</button>';
    }).join('');

    var rows = filtered.length === 0
      ? '<tr><td colspan="7" style="padding:48px;text-align:center;color:#aaa;font-size:14px">No transactions match your filters.</td></tr>'
      : filtered.map(function(tx){
          var upd  = state.updating[tx.id];
          var rBg  = tx.include ? '#f0faf4' : '#fff8f2';
          var bBg  = upd ? '#ddd' : (tx.include ? '#27ae60' : '#e67e22');
          var aCls = tx.amount >= 0 ? 'color:#27ae60' : 'color:#e74c3c';
          var typBg    = tx.type === 'income' ? '#e8f5ed' : '#fdecea';
          var typColor = tx.type === 'income' ? '#1c7a40' : '#c0392b';
          return '<tr style="background:' + rBg + ';border-bottom:1px solid #f0ede8">' +
            '<td style="padding:11px 14px;font-family:monospace;font-size:11px;color:#999;white-space:nowrap">' + h(tx.date) + '</td>' +
            '<td style="padding:11px 14px;font-size:12px;color:#1a1a1a;max-width:220px"><span class="clamp2">' + h(tx.description) + '</span></td>' +
            '<td style="padding:11px 14px;text-align:right;font-weight:700;font-size:13px;white-space:nowrap;' + aCls + '">' + (tx.amount>=0?'+':'') + fmt(tx.amount) + '</td>' +
            '<td style="padding:11px 14px;text-align:center"><span style="font-size:10px;font-weight:600;padding:3px 8px;background:' + typBg + ';color:' + typColor + '">' + h(tx.type) + '</span></td>' +
            '<td style="padding:11px 14px;font-size:11px;color:#666">' + h(tx.category) + '</td>' +
            '<td style="padding:11px 14px;text-align:center"><button data-toggle="' + h(tx.id) + '" ' + (upd?'disabled':'') + ' style="background:' + bBg + ';color:#fff;font-weight:700;font-size:11px;border:none;padding:5px 10px;cursor:pointer;min-width:36px">' +
              (upd ? '<svg class="spin" style="width:12px;height:12px" viewBox="0 0 24 24" fill="none"><circle style="opacity:.25" cx="12" cy="12" r="10" stroke="currentColor" stroke-width="4"/><path style="opacity:.75" fill="currentColor" d="M4 12a8 8 0 018-8V0C5.373 0 0 5.373 0 12h4z"/></svg>' : (tx.include?'Y':'N')) +
            '</button></td>' +
            '<td style="padding:11px 14px;font-size:11px;color:#aaa;max-width:200px"><span class="clamp2">' + h(tx.reason) + '</span></td>' +
          '</tr>';
        }).join('');

    return (
      '<div style="min-height:100vh;background:#f5f2ec;display:flex;flex-direction:column">' +
        '<div class="sc-hero-sm" style="display:flex;justify-content:space-between;align-items:center">' +
          '<div>' +
            '<div class="sc-serif" style="font-size:20px;color:#fff">Sahara Capital \u2014 Income Analyzer</div>' +
            '<div style="font-size:11px;color:rgba(255,255,255,0.5);margin-top:2px">Review and toggle transactions for mortgage underwriting</div>' +
          '</div>' +
          '<div style="display:flex;align-items:center;gap:20px">' +
            '<span class="sc-serif" style="font-size:13px;color:#c4a050;font-style:italic">by your side, always</span>' +
            '<img src="https://saharacapital.ca/wp-content/uploads/2026/01/Logo-02.png" alt="Sahara Capital" style="height:38px;width:auto">' +
          '</div>' +
        '</div>' +
        '<div class="sc-strip">' +
          '<div class="sc-strip-cell"><div class="sc-strip-num">' + totals.includedCount + ' / ' + totals.count + '</div><div class="sc-strip-lbl">Transactions Included</div></div>' +
          '<div class="sc-strip-cell"><div class="sc-strip-num">' + fmt(totals.income) + '</div><div class="sc-strip-lbl">Total Annual Income</div></div>' +
          '<div class="sc-strip-cell"><div class="sc-strip-num">' + fmt(totals.expenses) + '</div><div class="sc-strip-lbl">Total Annual Expenses</div></div>' +
          '<div class="sc-strip-cell"><div class="sc-strip-num" style="color:' + (totals.net>=0?'#c4a050':'#e74c3c') + '">' + fmt(totals.net) + '</div><div class="sc-strip-lbl">Net Annual Income</div></div>' +
        '</div>' +
        '<div style="background:#fff;border-bottom:1px solid #e8e4dc;padding:12px 20px">' +
          '<div style="max-width:1200px;margin:0 auto;display:flex;flex-wrap:wrap;gap:10px;align-items:center;justify-content:space-between">' +
            '<div style="display:flex;gap:4px;flex-wrap:wrap">' + filterBtns + '</div>' +
            '<div style="display:flex;align-items:center;gap:8px">' +
              '<input id="search-input" type="search" placeholder="Search transactions..." value="' + h(state.search) + '" style="padding:8px 14px;font-size:12px;border:1px solid #ddd;background:#faf9f7;outline:none;width:210px">' +
              (state.search.trim()
                ? '<button id="bulk-y-btn" style="background:#1B3D2C;color:#fff;font-size:11px;font-weight:700;border:none;padding:8px 14px;cursor:pointer">Set all Y</button>' +
                  '<button id="bulk-n-btn" style="background:#e67e22;color:#fff;font-size:11px;font-weight:700;border:none;padding:8px 14px;cursor:pointer">Set all N</button>'
                : '') +
            '</div>' +
          '</div>' +
        '</div>' +
        '<div style="flex:1;padding:20px;overflow:auto">' +
          '<div style="max-width:1200px;margin:0 auto">' +
            '<p style="font-size:11px;color:#aaa;margin-bottom:10px">Showing ' + filtered.length + ' of ' + state.transactions.length + ' transactions</p>' +
            '<div style="background:#fff;overflow:hidden">' +
              '<div style="overflow-x:auto">' +
                '<table style="width:100%;border-collapse:collapse">' +
                  '<thead><tr style="background:#1B3D2C">' +
                    '<th style="padding:11px 14px;text-align:left;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:rgba(255,255,255,0.7);white-space:nowrap">Date</th>' +
                    '<th style="padding:11px 14px;text-align:left;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:rgba(255,255,255,0.7)">Description</th>' +
                    '<th style="padding:11px 14px;text-align:right;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:rgba(255,255,255,0.7)">Amount</th>' +
                    '<th style="padding:11px 14px;text-align:center;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:rgba(255,255,255,0.7)">Type</th>' +
                    '<th style="padding:11px 14px;text-align:left;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:rgba(255,255,255,0.7)">Category</th>' +
                    '<th style="padding:11px 14px;text-align:center;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:rgba(255,255,255,0.7)">Include</th>' +
                    '<th style="padding:11px 14px;text-align:left;font-size:10px;font-weight:600;letter-spacing:1px;text-transform:uppercase;color:rgba(255,255,255,0.7)">Reason</th>' +
                  '</tr></thead>' +
                  '<tbody>' + rows + '</tbody>' +
                '</table>' +
              '</div>' +
            '</div>' +
          '</div>' +
        '</div>' +
        '<div style="background:#fff;border-top:1px solid #e8e4dc;padding:16px 24px">' +
          '<div style="max-width:1200px;margin:0 auto;display:flex;justify-content:flex-end">' +
            '<button id="to-download-btn" class="sc-btn sc-btn-gold" style="font-size:12px;padding:13px 28px">' +
              '<svg style="width:16px;height:16px" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"/></svg>' +
              'Generate Excel Report' +
            '</button>' +
          '</div>' +
        '</div>' +
      '</div>'
    );
  }

  /* ============================================================
     DOWNLOAD PAGE
  ============================================================ */
  function renderDownload() {
    var inc          = state.transactions.filter(function(t){return t.include;});
    var totalIncome  = inc.filter(function(t){return t.type==='income';}).reduce(function(s,t){return s+t.amount;},0);
    var totalExpenses= inc.filter(function(t){return t.type==='expense';}).reduce(function(s,t){return s+Math.abs(t.amount);},0);
    var net          = totalIncome - totalExpenses;
    var dlBtnBg      = state.downloaded ? '#1B3D2C' : '#c4a050';

    return (
      '<div class="sc-hero">' +
        '<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:30px">' +
          '<div class="sc-badge">Analysis Complete</div>' +
          '<img src="https://saharacapital.ca/wp-content/uploads/2026/01/Logo-02.png" alt="Sahara Capital" style="height:52px;width:auto">' +
        '</div>' +
        '<div style="font-size:10px;font-weight:600;letter-spacing:2px;color:rgba(255,255,255,0.5);text-transform:uppercase;margin-bottom:10px">Team Harman Batta</div>' +
        '<div class="sc-serif" style="font-size:42px;color:#ffffff;line-height:1.15">Sahara Capital Group</div>' +
        '<div class="sc-serif" style="font-size:26px;color:#c4a050;font-style:italic;margin-top:6px">by your side, always</div>' +
        '<div class="sc-divider" style="margin:20px 0"></div>' +
        '<div style="font-size:14px;color:rgba(255,255,255,0.6)">' + inc.length + ' of ' + state.transactions.length + ' transactions included in report</div>' +
      '</div>' +
      '<div class="sc-strip">' +
        '<div class="sc-strip-cell"><div class="sc-strip-num">' + fmt(totalIncome) + '</div><div class="sc-strip-lbl">Total Annual Income</div></div>' +
        '<div class="sc-strip-cell"><div class="sc-strip-num">' + fmt(totalExpenses) + '</div><div class="sc-strip-lbl">Total Annual Expenses</div></div>' +
        '<div class="sc-strip-cell"><div class="sc-strip-num" style="color:' + (net>=0?'#c4a050':'#e74c3c') + '">' + fmt(net) + '</div><div class="sc-strip-lbl">Net Annual Income</div></div>' +
      '</div>' +
      '<div style="background:#e8e4dc;padding:32px">' +
        '<div style="max-width:540px;margin:0 auto">' +
          '<div class="sc-card" style="background:#fff;padding:28px 32px;margin-bottom:20px">' +
            '<div class="sc-serif" style="font-size:18px;color:#1B3D2C;padding-bottom:12px;border-bottom:2px solid #c4a050;margin-bottom:16px">What\u2019s in the Excel Report</div>' +
            '<ul style="display:flex;flex-direction:column;gap:10px;font-size:13px;color:#555">' +
              '<li style="display:flex;gap:10px"><span style="color:#1B3D2C;font-weight:700">\u2713</span> Sheet 1: Summary \u2014 monthly income &amp; expense tables, annual totals, key metrics</li>' +
              '<li style="display:flex;gap:10px"><span style="color:#1B3D2C;font-weight:700">\u2713</span> Sheet 2: Income Breakdown \u2014 every deposit with Y/N colour coding</li>' +
              '<li style="display:flex;gap:10px"><span style="color:#1B3D2C;font-weight:700">\u2713</span> Sheet 3: Expense Breakdown \u2014 every withdrawal with Y/N colour coding</li>' +
              '<li style="display:flex;gap:10px"><span style="color:#1B3D2C;font-weight:700">\u2713</span> AI-generated category and reason for every transaction</li>' +
            '</ul>' +
          '</div>' +
          '<button id="download-btn" ' + (state.downloading?'disabled':'') + ' class="sc-btn sc-btn-full" style="font-size:13px;padding:16px 28px;margin-bottom:12px;background:' + dlBtnBg + (state.downloading?';opacity:.6;cursor:not-allowed':'') + '">' +
            (state.downloading
              ? '<svg class="spin" style="width:18px;height:18px" viewBox="0 0 24 24" fill="none"><circle style="opacity:.25" cx="12" cy="12" r="10" stroke="currentColor" stroke-width="4"/><path style="opacity:.75" fill="currentColor" d="M4 12a8 8 0 018-8V0C5.373 0 0 5.373 0 12h4z"/></svg>Generating Excel file...'
              : state.downloaded
                ? '\u2713 Downloaded! Click to download again'
                : '<svg style="width:18px;height:18px" fill="none" viewBox="0 0 24 24" stroke="currentColor"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"/></svg>Download Excel Report (.xlsx)') +
          '</button>' +
          '<div style="display:flex;gap:10px">' +
            '<button id="back-btn" style="flex:1;padding:13px;font-size:12px;font-weight:700;letter-spacing:1px;text-transform:uppercase;background:#fff;border:1px solid #c5bfb5;color:#666;cursor:pointer">\u2190 Back to Review</button>' +
            '<button id="reset-btn" style="flex:1;padding:13px;font-size:12px;font-weight:700;letter-spacing:1px;text-transform:uppercase;background:#ede9e1;border:none;color:#666;cursor:pointer">Analyze New Statements</button>' +
          '</div>' +
        '</div>' +
      '</div>' +
      '<div class="sc-contact">' +
        '<div class="sc-contact-info">' +
          '<div class="sc-serif" style="font-size:26px;color:#fff;margin-bottom:4px">Harman Batta</div>' +
          '<div style="font-size:10px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#c4a050;margin-bottom:4px">Mortgage Broker</div>' +
          '<div style="font-size:14px;color:rgba(255,255,255,0.7);margin-bottom:16px">Sahara Capital Group</div>' +
          '<div class="sc-serif" style="font-size:22px;color:#fff;margin-bottom:8px">647-685-9000</div>' +
          '<a href="mailto:info@saharacapital.ca" style="font-size:13px;color:rgba(255,255,255,0.6);display:block;text-decoration:none;margin-bottom:2px">info@saharacapital.ca</a>' +
          '<a href="https://saharacapital.ca" style="font-size:13px;color:rgba(255,255,255,0.6);display:block;text-decoration:none">saharacapital.ca</a>' +
          '<div style="font-size:11px;color:rgba(255,255,255,0.3);margin-top:8px">Serving all of Canada</div>' +
        '</div>' +
        '<div class="sc-contact-cta">' +
          '<div style="font-size:13px;color:#1B3D2C;font-weight:600;line-height:1.5;margin-bottom:16px">Powered by<br>Sahara Capital Group</div>' +
          '<a href="https://saharacapital.ca" style="display:block;background:#1B3D2C;color:#fff;text-decoration:none;font-size:11px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;padding:13px 20px;text-align:center;margin-bottom:12px">saharacapital.ca</a>' +
          '<div style="font-size:11px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#1B3D2C;margin-top:10px">Mortgage Alliance</div>' +
          '<div style="font-size:10px;color:rgba(27,61,44,0.6);margin-top:3px">FSRA Lic. #10530</div>' +
        '</div>' +
      '</div>' +
      '<div style="background:#f5f2ec;padding:12px 32px;text-align:center">' +
        '<p style="font-size:10.5px;color:#999">This tool is for internal mortgage underwriting use only. Sahara Capital Group &nbsp;|&nbsp; saharacapital.ca</p>' +
      '</div>'
    );
  }

  /* ============================================================
     EVENT LISTENERS
  ============================================================ */
  function attachListeners() {
    if      (state.page === 'upload')   attachUpload();
    else if (state.page === 'progress') attachProgress();
    else if (state.page === 'review')   attachReview();
    else                                attachDownload();
  }

  function attachUpload() {
    dragCounter = 0;   // reset counter whenever upload page re-renders
    var dz = document.getElementById('drop-zone');
    var fi = document.getElementById('file-input');
    var ab = document.getElementById('analyze-btn');

    if (dz) {
      // dragenter/dragleave counter prevents jitter when mouse passes over child elements
      dz.addEventListener('dragenter', function(e){
        e.preventDefault();
        dragCounter++;
        if (!state.isDragging) setState({isDragging: true});
      });
      dz.addEventListener('dragleave', function(e){
        dragCounter--;
        if (dragCounter <= 0) {
          dragCounter = 0;
          if (state.isDragging) setState({isDragging: false});
        }
      });
      dz.addEventListener('dragover', function(e){ e.preventDefault(); });
      dz.addEventListener('drop', function(e){
        e.preventDefault();
        dragCounter = 0;
        setState({isDragging: false});
        addFiles(e.dataTransfer.files);
      });
    }

    // File input change — fires after user picks files via the label button or drop
    if (fi) fi.addEventListener('change', function(e){
      addFiles(e.target.files);
      // Reset so the same file can be re-selected if removed then re-added
      fi.value = '';
    });

    // Remove buttons
    document.querySelectorAll('[data-remove]').forEach(function(btn){
      btn.addEventListener('click', function(e){
        e.stopPropagation();
        var idx = parseInt(btn.getAttribute('data-remove'));
        setState({files: state.files.filter(function(_,i){return i!==idx;})});
      });
    });

    if (ab) ab.addEventListener('click', handleAnalyze);
  }

  function attachProgress() {
    var rb = document.getElementById('retry-btn');
    if (rb) rb.addEventListener('click', function(){
      stopPolling();
      setState({page:'upload', jobId:null, progressInfo:{}, analyzeError:null});
    });
  }

  function attachReview() {
    document.querySelectorAll('[data-filter]').forEach(function(btn){
      btn.addEventListener('click', function(){ setState({filter: btn.getAttribute('data-filter')}); });
    });
    var si = document.getElementById('search-input');
    if (si) si.addEventListener('input', function(e){ setState({search: e.target.value}); });
    document.querySelectorAll('[data-toggle]').forEach(function(btn){
      btn.addEventListener('click', function(){ handleToggle(btn.getAttribute('data-toggle')); });
    });
    var byBtn = document.getElementById('bulk-y-btn');
    var bnBtn = document.getElementById('bulk-n-btn');
    if (byBtn) byBtn.addEventListener('click', function(){ handleBulk(true); });
    if (bnBtn) bnBtn.addEventListener('click', function(){ handleBulk(false); });
    var td = document.getElementById('to-download-btn');
    if (td) td.addEventListener('click', function(){ setState({page:'download', downloaded:false}); });
  }

  function attachDownload() {
    var db = document.getElementById('download-btn');
    if (db) db.addEventListener('click', handleDownload);
    var bb = document.getElementById('back-btn');
    if (bb) bb.addEventListener('click', function(){ setState({page:'review'}); });
    var rb = document.getElementById('reset-btn');
    if (rb) rb.addEventListener('click', function(){
      stopPolling();
      setState({page:'upload',files:[],sessionId:null,transactions:[],
                filter:'all',search:'',analyzeError:null,downloaded:false,
                jobId:null,progressInfo:{}});
    });
  }

  /* ============================================================
     ACTIONS
  ============================================================ */
  function addFiles(list) {
    if (!list) return;
    var pdfs = Array.from(list).filter(function(f){
      return f.type==='application/pdf' || f.name.toLowerCase().endsWith('.pdf');
    });
    var existing = {};
    state.files.forEach(function(f){ existing[f.name+f.size]=1; });
    var fresh = pdfs.filter(function(f){ return !existing[f.name+f.size]; });
    if (fresh.length) setState({files: state.files.concat(fresh), analyzeError:null});
  }

  async function handleAnalyze() {
    if (state.files.length===0) { setState({analyzeError:'Please add at least one PDF file.'}); return; }
    setState({isAnalyzing:true, analyzeError:null});
    try {
      var fd = new FormData();
      state.files.forEach(function(f){ fd.append('files',f); });
      var res = await fetch('/analyze',{method:'POST',body:fd});
      if (!res.ok) {
        var err = await res.json().catch(function(){return {detail:'Unknown error'};});
        throw new Error(err.detail||'Error '+res.status);
      }
      var data = await res.json();
      setState({
        isAnalyzing: false,
        jobId: data.job_id,
        page: 'progress',
        progressInfo: {status:'processing', current_file:'Starting\u2026', pages_done:0, total_pages:0}
      });
      startPolling(data.job_id);
    } catch(e) {
      setState({isAnalyzing:false, analyzeError:e.message||'Failed to analyze. Please try again.'});
    }
  }

  function startPolling(jobId) {
    stopPolling();
    setTimeout(function(){ pollJob(jobId); }, 2000);
    pollTimer = setInterval(function(){ pollJob(jobId); }, 5000);
  }

  function stopPolling() {
    if (pollTimer) { clearInterval(pollTimer); pollTimer = null; }
  }

  async function pollJob(jobId) {
    if (state.page !== 'progress') { stopPolling(); return; }
    try {
      var res = await fetch('/jobs/'+jobId+'/status');
      if (!res.ok) return;
      var data = await res.json();
      if (data.status === 'complete') {
        stopPolling();
        var txRes = await fetch('/sessions/'+data.session_id+'/transactions');
        if (!txRes.ok) {
          setState({progressInfo:{status:'error', error:'Failed to load transactions.', current_file:'', pages_done:0, total_pages:0}});
          return;
        }
        var transactions = await txRes.json();
        setState({page:'review', sessionId:data.session_id, transactions:transactions,
                  progressInfo:data, filter:'all', search:''});
      } else if (data.status === 'error') {
        stopPolling();
        setState({progressInfo: data});
      } else {
        setState({progressInfo: data});
      }
    } catch(e) { /* silent — retries on next interval */ }
  }

  async function handleToggle(txId) {
    var tx = state.transactions.find(function(t){return t.id===txId;});
    if (!tx || state.updating[txId]) return;
    var newInclude = !tx.include;

    var optimisticTxs = state.transactions.map(function(t){
      return t.id===txId ? Object.assign({},t,{include:newInclude}) : t;
    });
    var upd = Object.assign({},state.updating); upd[txId]=true;
    setState({transactions:optimisticTxs, updating:upd});

    try {
      var res = await fetch('/sessions/'+state.sessionId+'/transactions/'+txId,{
        method:'PUT', headers:{'Content-Type':'application/json'},
        body:JSON.stringify({include:newInclude})
      });
      if (!res.ok) throw new Error('Failed');
      var updated = await res.json();
      var txs = state.transactions.map(function(t){return t.id===updated.id?updated:t;});
      var done = Object.assign({},state.updating); delete done[txId];
      setState({transactions:txs, updating:done});
    } catch(e) {
      var revertTxs = state.transactions.map(function(t){
        return t.id===txId ? Object.assign({},t,{include:!newInclude}) : t;
      });
      var done2 = Object.assign({},state.updating); delete done2[txId];
      setState({transactions:revertTxs, updating:done2});
    }
  }

  async function handleBulk(include) {
    if (!state.search.trim() || !state.sessionId) return;
    var q = state.search.trim();
    var label = include ? 'Y' : 'N';
    if (!confirm('Set all transactions matching "' + q + '" to ' + label + '?')) return;

    // Optimistic update
    var updated = state.transactions.map(function(t){
      return q && t.description.toLowerCase().indexOf(q.toLowerCase()) >= 0
        ? Object.assign({}, t, {include: include})
        : t;
    });
    setState({transactions: updated});

    try {
      await fetch('/sessions/' + state.sessionId + '/transactions/bulk', {
        method: 'PUT',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({description: q, include: include})
      });
    } catch(e) {
      console.error('Bulk update failed', e);
    }
  }

  async function handleDownload() {
    setState({downloading:true});
    try {
      var res = await fetch('/sessions/'+state.sessionId+'/export');
      if (!res.ok) throw new Error('Failed to generate export');
      var blob = await res.blob();
      var url  = URL.createObjectURL(blob);
      var a    = document.createElement('a');
      a.href=url; a.download='bank_statement_analysis_'+new Date().toISOString().slice(0,10)+'.xlsx';
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
      URL.revokeObjectURL(url);
      setState({downloading:false, downloaded:true});
    } catch(e) {
      setState({downloading:false});
      alert('Failed to download. Please try again.');
    }
  }

  render();
})();
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# FastAPI app
# ─────────────────────────────────────────────────────────────────────────────

app = FastAPI(title="Mortgage Bank Statement Analyzer")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])


@app.get("/")
def index():
    return HTMLResponse(HTML)


@app.get("/health")
def health():
    return {"status": "ok"}


# ─────────────────────────────────────────────────────────────────────────────
# Background job: Stage 1 (Python extraction) → Stage 2 (Claude categorisation)
# ─────────────────────────────────────────────────────────────────────────────

async def _process_job(job_id: str, file_data: list[tuple[str, bytes]]):
    """
    For each uploaded PDF:
      Stage 1 — Python parser extracts exact structured transactions (no AI).
      Stage 2 — Claude categorises in batches of 150; only assigns category,
                 suggested_include, and reason — direction/amount already known.
      Stage 3 — Python dedup pass normalises any cross-batch inconsistencies.
    """
    BATCH_SIZE   = 150
    all_transactions: list[dict] = []
    total_files = len(file_data)

    try:
        for file_idx, (filename, pdf_bytes) in enumerate(file_data):
            jobs[job_id].update({
                "current_file": f"File {file_idx+1}/{total_files}: {filename} — extracting transactions…",
                "pages_done":   0,
                "total_pages":  0,
            })

            # ── Stage 1a: Try Python text extraction ──────────────────────────
            raw_txs = await asyncio.to_thread(
                extract_transactions_from_pdf, pdf_bytes, filename
            )

            # Extract account holder name for same-name transfer detection
            account_holder = await asyncio.to_thread(
                _extract_account_holder, pdf_bytes
            )
            if account_holder:
                print(f"[INFO] {filename}: account holder detected — '{account_holder}'")

            # Build account-holder context injected into both prompts
            acct_context = ""
            if account_holder:
                acct_context = (
                    f"\n\nACCOUNT HOLDER NAME: \"{account_holder}\"\n"
                    f"SAME-NAME RULE: Any e-transfer sent TO or received FROM a name that "
                    f"matches or closely resembles \"{account_holder}\" is an internal transfer "
                    f"between the account holder's own accounts (e.g. chequing ↔ savings). "
                    f"Set suggested_include N, category \"Internal Transfer\", "
                    f"reason \"Internal transfer between own accounts — excluded.\" "
                    f"This applies to BOTH the income side (received from own name) AND the "
                    f"expense side (sent to own name). Never set Y for these.\n"
                )

            # ── Stage 1b: Fallback to Claude vision for image-based PDFs ─────
            if not raw_txs:
                jobs[job_id]["current_file"] = (
                    f"File {file_idx+1}/{total_files}: {filename} — "
                    f"⚠️ Scanned or image-based PDF detected. Reading with vision…"
                )
                vision_txs = await _extract_via_vision(
                    pdf_bytes, filename, extra_context=acct_context
                )

                if not vision_txs:
                    print(f"[WARN] {filename}: could not extract transactions "
                          f"via text or vision — skipping")
                    continue

                # Vision transactions are already categorised — skip Stage 2
                all_transactions.extend(vision_txs)
                jobs[job_id].update({
                    "pages_done":  1,
                    "total_pages": 1,
                    "current_file": (
                        f"File {file_idx+1}/{total_files}: {filename} — "
                        f"⚠️ Scanned PDF: {len(vision_txs)} transactions extracted via vision. "
                        f"For higher accuracy, ask your client to re-download this statement "
                        f"directly from their online banking portal."
                    ),
                })
                print(f"[INFO] {filename}: vision path complete — "
                      f"{len(vision_txs)} transactions")
                continue

            # ── Stage 2: Claude categorisation (text-based PDFs only) ─────────
            print(f"[INFO] {filename}: {len(raw_txs)} transactions ready for categorisation")
            total_batches = (len(raw_txs) + BATCH_SIZE - 1) // BATCH_SIZE
            jobs[job_id]["total_pages"] = total_batches

            categorized: dict[str, dict] = {}

            for batch_num, start in enumerate(range(0, len(raw_txs), BATCH_SIZE), 1):
                batch = raw_txs[start: start + BATCH_SIZE]
                jobs[job_id]["current_file"] = (
                    f"File {file_idx+1}/{total_files}: {filename} — "
                    f"categorising batch {batch_num}/{total_batches}…"
                )

                batch_input = [
                    {
                        "id":          tx["id"],
                        "description": tx["description"],
                        "amount":      tx["amount"],
                        "direction":   tx["direction"],
                        "date":        tx["date"],
                        "month":       tx["month"],
                    }
                    for tx in batch
                ]

                prompt = CAT_PROMPT + acct_context + "\n\nTransactions to categorise:\n" + json.dumps(batch_input)

                def _call(p=prompt) -> str:
                    with client.messages.stream(
                        model="claude-haiku-4-5-20251001",
                        max_tokens=16000,
                        messages=[{"role": "user", "content": p}],
                    ) as stream:
                        return stream.get_final_text().strip()

                try:
                    response_text = await asyncio.to_thread(_call)
                    results       = extract_json_transactions(response_text)
                    for item in results:
                        if "id" in item:
                            categorized[item["id"]] = item
                    print(f"[INFO] {filename} batch {batch_num}/{total_batches}: "
                          f"{len(results)} categorised")
                except Exception as e:
                    print(f"[WARN] Claude error — {filename} batch {batch_num}: {e}")

                jobs[job_id]["pages_done"] = batch_num

            # ── Stage 3: Merge categorisation with extracted transactions ─────
            file_transactions: list[dict] = []
            for tx in raw_txs:
                cat      = categorized.get(tx["id"], {})
                tx_type  = "income" if tx["direction"] == "credit" else "expense"
                inc_flag = str(cat.get("suggested_include", "Y")).strip().upper() == "Y"
                signed   = abs(tx["amount"]) if tx_type == "income" else -abs(tx["amount"])

                file_transactions.append({
                    "id":               tx["id"],
                    "date":             tx["date"],
                    "description":      tx["description"],
                    "amount":           signed,
                    "type":             tx_type,
                    "month":            tx["month"],
                    "category":         cat.get("category", "Other"),
                    "suggested_include": cat.get("suggested_include", "Y"),
                    "reason":           cat.get("reason", ""),
                    "include":          inc_flag,
                })

            all_transactions.extend(file_transactions)
            print(f"[INFO] {filename}: complete — {len(file_transactions)} transactions merged")

        # Dedup categories across all files for cross-batch consistency
        if all_transactions:
            # Strip internal _vision flag before storing
            for tx in all_transactions:
                tx.pop("_vision", None)
            all_transactions = deduplicate_categories(all_transactions)

        session_id = str(uuid.uuid4())
        sessions[session_id] = {"transactions": all_transactions}
        jobs[job_id].update({
            "status":            "complete",
            "session_id":        session_id,
            "transaction_count": len(all_transactions),
            "current_file":      f"Complete — {len(all_transactions)} transactions from {total_files} file(s)",
        })
        print(f"[INFO] Job {job_id} complete: {len(all_transactions)} transactions, "
              f"session {session_id}")

    except Exception as e:
        jobs[job_id] = {
            "status":      "error",
            "error":       str(e),
            "current_file": "",
            "pages_done":  0,
            "total_pages": 0,
        }
        print(f"[ERROR] Job {job_id} failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# API routes
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/analyze")
async def analyze_statements(files: list[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"{file.filename} is not a PDF")

    file_data = [(f.filename, await f.read()) for f in files]
    job_id    = str(uuid.uuid4())
    jobs[job_id] = {
        "status":      "processing",
        "current_file": f"Starting — 0 of {len(file_data)} files processed",
        "pages_done":  0,
        "total_pages": 0,
    }

    task = asyncio.create_task(_process_job(job_id, file_data))
    _running_tasks.add(task)
    task.add_done_callback(_running_tasks.discard)

    return {"job_id": job_id}


@app.get("/jobs/{job_id}/status")
def get_job_status(job_id: str):
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = jobs[job_id]
    resp = {
        "status":       job["status"],
        "current_file": job.get("current_file", ""),
        "pages_done":   job.get("pages_done", 0),
        "total_pages":  job.get("total_pages", 0),
    }
    if job["status"] == "complete":
        resp["session_id"]        = job["session_id"]
        resp["transaction_count"] = job["transaction_count"]
    elif job["status"] == "error":
        resp["error"] = job.get("error", "Unknown error")
    return resp


@app.get("/sessions/{session_id}/transactions")
def get_session_transactions(session_id: str):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    return sessions[session_id]["transactions"]


@app.put("/sessions/{session_id}/transactions/{transaction_id}")
def update_transaction(session_id: str, transaction_id: str, body: UpdateTransactionRequest):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    for tx in sessions[session_id]["transactions"]:
        if tx["id"] == transaction_id:
            tx["include"] = body.include
            if body.reason is not None:
                tx["reason"] = body.reason
            return tx
    raise HTTPException(status_code=404, detail="Transaction not found")


@app.put("/sessions/{session_id}/transactions/bulk")
def bulk_update_transactions(session_id: str, body: BulkUpdateRequest):
    """Set include=Y or N for all transactions whose description matches the given substring."""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")
    q = body.description.lower()
    updated = 0
    for tx in sessions[session_id]["transactions"]:
        if q in tx.get("description", "").lower():
            tx["include"] = body.include
            updated += 1
    return {"updated": updated}


@app.get("/sessions/{session_id}/export")
def export_excel(session_id: str):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    transactions = sessions[session_id]["transactions"]
    income_txs   = [tx for tx in transactions if tx.get("type") == "income"]
    expense_txs  = [tx for tx in transactions if tx.get("type") == "expense"]

    def parse_month(m):
        try:    return datetime.strptime(m, "%b-%y")
        except: return datetime.min

    all_months     = sorted(set(tx.get("month","") for tx in transactions if tx.get("month")), key=parse_month)
    income_months  = sorted(set(tx.get("month","") for tx in income_txs  if tx.get("month")), key=parse_month)
    expense_months = sorted(set(tx.get("month","") for tx in expense_txs if tx.get("month")), key=parse_month)
    income_cats    = sorted(set(tx.get("category","") for tx in income_txs  if tx.get("category")))
    expense_cats   = sorted(set(tx.get("category","") for tx in expense_txs if tx.get("category")))

    n_inc = len(income_months)  or 1
    n_exp = len(expense_months) or 1
    n_all = len(all_months)     or 1

    # ── Styles ────────────────────────────────────────────────────────────────
    green_row  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    orange_row = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    inc_y_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    inc_n_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    hdr_fill   = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    sec_fill   = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    sub_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    tot_fill   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    met_fill   = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    net_pos    = PatternFill(start_color="E8F8E8", end_color="E8F8E8", fill_type="solid")

    hdr_font  = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    sec_font  = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    bold_font = Font(name="Arial", bold=True, size=10)
    norm_font = Font(name="Arial", size=10)
    wb_font   = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    net_p_fnt = Font(name="Arial", bold=True, size=10, color="27AE60")
    net_n_fnt = Font(name="Arial", bold=True, size=10, color="E74C3C")

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    mfmt = "#,##0.00"
    pfmt = "0.0%"
    MIN_H = 20

    # ── Helpers ───────────────────────────────────────────────────────────────
    def autofit_columns(ws, min_width=8, max_width=60):
        col_widths: dict[int, int] = {}
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                try:
                    curr = cell.alignment if cell.alignment else Alignment()
                    cell.alignment = Alignment(
                        horizontal=curr.horizontal, vertical=curr.vertical or "center",
                        wrap_text=True, indent=curr.indent,
                    )
                    if cell.value is not None and not str(cell.value).startswith("="):
                        col = cell.column
                        length = max(len(ln) for ln in str(cell.value).split("\n")) + 2
                        col_widths[col] = max(col_widths.get(col, min_width), length)
                except Exception:
                    pass
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = min(max(width, min_width), max_width)
        for ri in ws.row_dimensions:
            if (ws.row_dimensions[ri].height or 0) < MIN_H:
                ws.row_dimensions[ri].height = MIN_H

    def sec_hdr(ws, r, text, ncols):
        c = ws.cell(row=r, column=1, value=text)
        c.font = sec_font; c.fill = sec_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 22
        if ncols > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)

    def col_hdrs(ws, r, hdrs, start_col=1):
        for ci, h in enumerate(hdrs, start_col):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = hdr_font; c.fill = sub_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin
        ws.row_dimensions[r].height = MIN_H

    def num_cell(ws, r, ci, formula, font=None, fill=None):
        c = ws.cell(row=r, column=ci, value=formula)
        c.font = font or norm_font; c.border = thin; c.number_format = mfmt
        c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        if fill: c.fill = fill

    def lbl_cell(ws, r, ci, text, font=None, fill=None):
        c = ws.cell(row=r, column=ci, value=text)
        c.font = font or norm_font; c.border = thin
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if fill: c.fill = fill

    # ── Breakdown sheet constants ─────────────────────────────────────────────
    BD_DATA_START = 3
    BD_MAX_ROW    = 2000
    INC_SHEET     = "Income Breakdown"
    EXP_SHEET     = "Expense Breakdown"

    def sp(sheet, col, r1=BD_DATA_START, r2=BD_MAX_ROW):
        return f"'{sheet}'!${col}${r1}:${col}${r2}"

    def sumproduct_cat_month(sheet, cat, month):
        return (f"=SUMPRODUCT("
                f"({sp(sheet,'E')}=\"{cat}\")*"
                f"({sp(sheet,'D')}=\"{month}\")*"
                f"({sp(sheet,'F')}=\"Y\")*"
                f"{sp(sheet,'C')})")

    def sumproduct_cat_annual(sheet, cat):
        return (f"=SUMPRODUCT("
                f"({sp(sheet,'E')}=\"{cat}\")*"
                f"({sp(sheet,'F')}=\"Y\")*"
                f"{sp(sheet,'C')})")

    def sumproduct_cat_has_y(sheet, cat):
        return (f"=IF(SUMPRODUCT("
                f"({sp(sheet,'E')}=\"{cat}\")*"
                f"({sp(sheet,'F')}=\"Y\"))>0,\"Y\",\"N\")")

    # ── Summary sheet row layout ──────────────────────────────────────────────
    S1_HDR        = 1
    S1_COLHDR     = 2
    S1_DATA_START = 3
    S1_DATA_END   = S1_DATA_START + max(len(income_months),  1) - 1
    S1_GRAND      = S1_DATA_END + 1
    S1_AVG        = S1_GRAND + 1

    S2_HDR        = S1_AVG + 2
    S2_COLHDR     = S2_HDR + 1
    S2_DATA_START = S2_COLHDR + 1
    S2_DATA_END   = S2_DATA_START + max(len(expense_months), 1) - 1
    S2_GRAND      = S2_DATA_END + 1
    S2_AVG        = S2_GRAND + 1

    S3_HDR        = S2_AVG + 2
    S3_COLHDR     = S3_HDR + 1
    S3_DATA_START = S3_COLHDR + 1
    n_cat_rows    = max(len(income_cats), len(expense_cats), 1)
    S3_DATA_END   = S3_DATA_START + n_cat_rows - 1

    S4_HDR        = S3_DATA_END + 2
    S4_DATA_START = S4_HDR + 1

    inc_total_col = len(income_cats)  + 2
    exp_total_col = len(expense_cats) + 2
    max_data_cols = max(inc_total_col, exp_total_col, 9)

    inc_grand_ref = f"{get_column_letter(inc_total_col)}{S1_GRAND}"
    exp_grand_ref = f"{get_column_letter(exp_total_col)}{S2_GRAND}"

    # ── Build Workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Section 1 — Monthly Income
    sec_hdr(ws, S1_HDR, "SECTION 1 — MONTHLY INCOME BREAKDOWN  (Y transactions only)", inc_total_col)
    col_hdrs(ws, S1_COLHDR, ["Month"] + income_cats + ["Total Income"])
    for mi, mo in enumerate(income_months):
        r = S1_DATA_START + mi
        lbl_cell(ws, r, 1, mo)
        for ci, cat in enumerate(income_cats, 2):
            num_cell(ws, r, ci, sumproduct_cat_month(INC_SHEET, cat, mo))
        tot_f = f"=SUM({get_column_letter(2)}{r}:{get_column_letter(1+len(income_cats))}{r})" if income_cats else "=0"
        num_cell(ws, r, inc_total_col, tot_f)
        ws.row_dimensions[r].height = MIN_H
    if not income_months:
        ws.cell(row=S1_DATA_START, column=1, value="(no income data)").font = norm_font
    lbl_cell(ws, S1_GRAND, 1, "Grand Total", bold_font, tot_fill)
    for ci in range(2, inc_total_col + 1):
        cl = get_column_letter(ci)
        num_cell(ws, S1_GRAND, ci, f"=SUM({cl}{S1_DATA_START}:{cl}{S1_DATA_END})", bold_font, tot_fill)
    ws.row_dimensions[S1_GRAND].height = MIN_H
    lbl_cell(ws, S1_AVG, 1, "Monthly Average", bold_font, tot_fill)
    for ci in range(2, inc_total_col + 1):
        cl = get_column_letter(ci)
        num_cell(ws, S1_AVG, ci, f"={cl}{S1_GRAND}/{n_inc}", bold_font, tot_fill)
    ws.row_dimensions[S1_AVG].height = MIN_H

    # Section 2 — Monthly Expenses
    sec_hdr(ws, S2_HDR, "SECTION 2 — MONTHLY EXPENSE BREAKDOWN  (Y transactions only)", exp_total_col)
    col_hdrs(ws, S2_COLHDR, ["Month"] + expense_cats + ["Total Expenses"])
    for mi, mo in enumerate(expense_months):
        r = S2_DATA_START + mi
        lbl_cell(ws, r, 1, mo)
        for ci, cat in enumerate(expense_cats, 2):
            num_cell(ws, r, ci, sumproduct_cat_month(EXP_SHEET, cat, mo))
        tot_f = f"=SUM({get_column_letter(2)}{r}:{get_column_letter(1+len(expense_cats))}{r})" if expense_cats else "=0"
        num_cell(ws, r, exp_total_col, tot_f)
        ws.row_dimensions[r].height = MIN_H
    if not expense_months:
        ws.cell(row=S2_DATA_START, column=1, value="(no expense data)").font = norm_font
    lbl_cell(ws, S2_GRAND, 1, "Grand Total", bold_font, tot_fill)
    for ci in range(2, exp_total_col + 1):
        cl = get_column_letter(ci)
        num_cell(ws, S2_GRAND, ci, f"=SUM({cl}{S2_DATA_START}:{cl}{S2_DATA_END})", bold_font, tot_fill)
    ws.row_dimensions[S2_GRAND].height = MIN_H
    lbl_cell(ws, S2_AVG, 1, "Monthly Average", bold_font, tot_fill)
    for ci in range(2, exp_total_col + 1):
        cl = get_column_letter(ci)
        num_cell(ws, S2_AVG, ci, f"={cl}{S2_GRAND}/{n_exp}", bold_font, tot_fill)
    ws.row_dimensions[S2_AVG].height = MIN_H

    # Section 3 — Annual Category Totals
    sec_hdr(ws, S3_HDR, "SECTION 3 — ANNUAL CATEGORY TOTALS", max(max_data_cols, 9))
    col_hdrs(ws, S3_COLHDR, ["Category", "Annual Total", "% of Total Income",    "Include Status"], 1)
    col_hdrs(ws, S3_COLHDR, ["Category", "Annual Total", "% of Total Expenses",  "Include Status"], 6)
    for idx in range(n_cat_rows):
        r = S3_DATA_START + idx
        ws.row_dimensions[r].height = MIN_H
        if idx < len(income_cats):
            cat   = income_cats[idx]
            has_y = any(t.get("include", True) for t in income_txs if t.get("category") == cat)
            c1 = ws.cell(row=r, column=1, value=cat);  c1.font=norm_font; c1.border=thin; c1.alignment=Alignment(horizontal="left",  vertical="center", wrap_text=True)
            c2 = ws.cell(row=r, column=2, value=sumproduct_cat_annual(INC_SHEET, cat)); c2.font=norm_font; c2.border=thin; c2.number_format=mfmt; c2.alignment=Alignment(horizontal="right", vertical="center", wrap_text=True)
            c3 = ws.cell(row=r, column=3, value=f"=IF({inc_grand_ref}<>0,B{r}/{inc_grand_ref},0)"); c3.font=norm_font; c3.border=thin; c3.number_format=pfmt; c3.alignment=Alignment(horizontal="right", vertical="center", wrap_text=True)
            c4 = ws.cell(row=r, column=4, value=sumproduct_cat_has_y(INC_SHEET, cat)); c4.font=wb_font; c4.fill=inc_y_fill if has_y else inc_n_fill; c4.border=thin; c4.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        if idx < len(expense_cats):
            cat   = expense_cats[idx]
            has_y = any(t.get("include", True) for t in expense_txs if t.get("category") == cat)
            c6 = ws.cell(row=r, column=6, value=cat);  c6.font=norm_font; c6.border=thin; c6.alignment=Alignment(horizontal="left",  vertical="center", wrap_text=True)
            c7 = ws.cell(row=r, column=7, value=sumproduct_cat_annual(EXP_SHEET, cat)); c7.font=norm_font; c7.border=thin; c7.number_format=mfmt; c7.alignment=Alignment(horizontal="right", vertical="center", wrap_text=True)
            c8 = ws.cell(row=r, column=8, value=f"=IF({exp_grand_ref}<>0,G{r}/{exp_grand_ref},0)"); c8.font=norm_font; c8.border=thin; c8.number_format=pfmt; c8.alignment=Alignment(horizontal="right", vertical="center", wrap_text=True)
            c9 = ws.cell(row=r, column=9, value=sumproduct_cat_has_y(EXP_SHEET, cat)); c9.font=wb_font; c9.fill=inc_y_fill if has_y else inc_n_fill; c9.border=thin; c9.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rng in (f"D{S3_DATA_START}:D{S3_DATA_END}", f"I{S3_DATA_START}:I{S3_DATA_END}"):
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Y"'], fill=inc_y_fill, font=wb_font))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"N"'], fill=inc_n_fill, font=wb_font))

    # Section 4 — Key Metrics
    sec_hdr(ws, S4_HDR, "SECTION 4 — KEY METRICS", 4)
    metrics = [
        ("Total Annual Income",        f"={inc_grand_ref}",                          False),
        ("Total Annual Expenses",      f"={exp_grand_ref}",                          False),
        ("Net Annual Income",          f"={inc_grand_ref}-{exp_grand_ref}",          True),
        ("Monthly Average Income",     f"={inc_grand_ref}/{n_inc}",                  False),
        ("Monthly Average Expenses",   f"={exp_grand_ref}/{n_exp}",                  False),
        ("Monthly Average Net Income", f"=({inc_grand_ref}-{exp_grand_ref})/{n_all}",True),
    ]
    for i, (lbl, formula, _) in enumerate(metrics):
        r  = S4_DATA_START + i
        lbl_cell(ws, r, 1, lbl, bold_font, met_fill)
        vc = ws.cell(row=r, column=2, value=formula)
        vc.fill=met_fill; vc.border=thin; vc.number_format=mfmt; vc.font=bold_font
        vc.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = MIN_H
    for i in (2, 5):
        nc = f"B{S4_DATA_START + i}"
        ws.conditional_formatting.add(nc, CellIsRule(operator="greaterThanOrEqual", formula=["0"],  fill=net_pos,  font=net_p_fnt))
        ws.conditional_formatting.add(nc, CellIsRule(operator="lessThan",           formula=["0"],  fill=met_fill, font=net_n_fnt))

    autofit_columns(ws)

    # ── Branding footer (safe — appended after all existing content/formulas) ──
    brand_row = S4_DATA_START + len(metrics) + 2
    bc = ws.cell(row=brand_row, column=1,
                 value="Prepared by Team Harman Batta  |  Sahara Capital Group  |  Mortgage Alliance  |  FSRA Lic. #10530  |  saharacapital.ca")
    bc.font = Font(name='Arial', size=9, color='1B3D2C', italic=True)
    bc.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[brand_row].height = MIN_H

    # ── Breakdown sheets ──────────────────────────────────────────────────────
    def build_breakdown_sheet(ws_b, txs):
        banner = ("GREEN rows (Y) are counted in Summary totals.  "
                  "ORANGE rows (N) are excluded but visible for reference.  "
                  "Change the Include dropdown (F column) to Y or N — "
                  "the Summary sheet will recalculate automatically.")
        bc = ws_b.cell(row=1, column=1, value=banner)
        bc.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        bc.fill = hdr_fill
        bc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_b.merge_cells("A1:G1")
        ws_b.row_dimensions[1].height = 40

        for ci, h in enumerate(["Date","Description","Amount","Month","Category","Include","Notes / Reason"], 1):
            c = ws_b.cell(row=2, column=ci, value=h)
            c.font=hdr_font; c.fill=hdr_fill
            c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border=thin
        ws_b.row_dimensions[2].height = MIN_H

        sorted_txs = sorted(txs, key=lambda t: t.get("date",""))
        for ri, tx in enumerate(sorted_txs, BD_DATA_START):
            inc  = tx.get("include", True)
            rf   = green_row if inc else orange_row
            vals = [
                tx.get("date",""),
                tx.get("description",""),
                abs(tx.get("amount", 0)),
                tx.get("month",""),
                tx.get("category",""),
                "Y" if inc else "N",
                tx.get("reason",""),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws_b.cell(row=ri, column=ci, value=val)
                c.fill=rf; c.border=thin
                if ci == 3:
                    c.font=norm_font; c.number_format=mfmt
                    c.alignment=Alignment(horizontal="right", vertical="center", wrap_text=True)
                elif ci == 6:
                    c.font=wb_font; c.fill=inc_y_fill if inc else inc_n_fill
                    c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif ci == 7:
                    c.font=norm_font; c.alignment=Alignment(vertical="center", wrap_text=True)
                else:
                    c.font=norm_font; c.alignment=Alignment(vertical="center", wrap_text=True)
            ws_b.row_dimensions[ri].height = MIN_H

        dv_last = max(len(sorted_txs) + BD_DATA_START + 50, BD_MAX_ROW)
        dv = DataValidation(
            type="list", formula1='"Y,N"', allow_blank=False,
            showDropDown=False, showErrorMessage=True,
            error="Enter Y or N", errorTitle="Invalid value",
        )
        dv.sqref = f"F{BD_DATA_START}:F{dv_last}"
        ws_b.add_data_validation(dv)
        autofit_columns(ws_b)

    ws_inc = wb.create_sheet(INC_SHEET)
    build_breakdown_sheet(ws_inc, income_txs)
    ws_exp = wb.create_sheet(EXP_SHEET)
    build_breakdown_sheet(ws_exp, expense_txs)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)

    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"bank_statement_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
